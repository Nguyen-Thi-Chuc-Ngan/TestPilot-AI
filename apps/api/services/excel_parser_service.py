"""
Excel/CSV parser for test case files.
Handles messy column names, missing fields, and bad data gracefully.
"""
import io
import csv
import uuid
from typing import Optional
import openpyxl
import pandas as pd
import structlog

logger = structlog.get_logger()

# Column name normalization map — user's messy names → our canonical names
COLUMN_MAP = {
    # TC ID
    "tc_id": "tc_id", "#": "tc_id", "no": "tc_id", "no.": "tc_id", "id": "tc_id",
    "test id": "tc_id", "testid": "tc_id", "test case id": "tc_id",

    # Module/Feature
    "module": "module", "feature": "feature", "component": "feature",
    "function": "feature", "chức năng": "feature",

    # Test type
    "test type": "test_type", "type": "test_type", "loại test": "test_type",

    # Priority
    "priority": "priority", "ưu tiên": "priority", "p": "priority",

    # Severity
    "severity": "severity", "mức độ": "severity",

    # Precondition
    "precondition": "precondition", "preconditions": "precondition",
    "pre-condition": "precondition", "điều kiện tiên quyết": "precondition",
    "điều kiện": "precondition",

    # Description — many variations
    "test case description": "description", "description": "description",
    "test case": "description", "test description": "description",
    "mô tả": "description", "mô tả test case": "description",
    "title": "description", "name": "description",
    "test case name": "description", "case description": "description",
    "scenario": "description", "test scenario": "description",
    "kịch bản": "description", "tên test case": "description",

    # Steps
    "steps": "steps", "actions": "steps", "test steps": "steps",
    "step": "steps", "action": "steps", "test action": "steps",
    "các bước": "steps", "bước thực hiện": "steps",
    "test steps/actions": "steps", "steps/actions": "steps",

    # Test data
    "test data": "test_data", "data": "test_data", "input": "test_data",
    "dữ liệu test": "test_data", "test input": "test_data",
    "input data": "test_data", "test value": "test_data",

    # Expected result
    "expected result": "expected_result", "expected": "expected_result",
    "expected results": "expected_result", "kết quả mong đợi": "expected_result",
    "kq mong đợi": "expected_result", "kết quả kỳ vọng": "expected_result",
    "expected outcome": "expected_result", "result expected": "expected_result",

    # Actual result
    "actual result": "actual_result", "actual": "actual_result",
    "actual results": "actual_result", "kết quả thực tế": "actual_result",
    "kq thực tế": "actual_result",

    # Status — including common Vietnamese QA file variants
    "status": "status", "result": "status", "trạng thái": "status",
    "kết quả": "status", "status build": "status", "status b": "status",
    "pass/fail": "status", "test result": "status", "kq": "status",
    "status run": "status", "run status": "status",

    # Bug
    "bug id": "bug_id", "bug": "bug_id", "defect id": "bug_id",
    "defect": "bug_id", "jira": "bug_id", "fix": "bug_id",
    "bug ref": "bug_id", "issue": "bug_id", "ticket": "bug_id",

    # Environment
    "environment": "environment", "env": "environment", "môi trường": "environment",
    "browser": "browser", "platform": "platform",

    # Automation
    "automation status": "automation_status", "automated": "automation_status",
    "automation": "automation_status",

    # People
    "created by": "created_by", "executed by": "executed_by",
    "executor": "executed_by", "tester": "executed_by",
    "dev": "dev_owner", "dev owner": "dev_owner",
    "dev đã fix": "dev_fixed", "fixed": "dev_fixed", "dev fixed": "dev_fixed",
    "dev fix": "dev_fixed",

    # Date
    "execution date": "execution_date", "date": "execution_date",
    "ngày thực hiện": "execution_date",

    # Notes
    "notes": "notes", "note": "notes", "comment": "notes",
    "ghi chú": "notes", "nhận xét": "notes", "remarks": "notes",
    "observation": "notes", "comment/notes": "notes",

    # Tags
    "tags": "tags", "tag": "tags", "label": "tags",

    # Requirement
    "requirement id": "requirement_id", "req id": "requirement_id",
    "requirement": "requirement_id", "user story": "requirement_id",
}

STATUS_MAP = {
    "pass": "Passed", "passed": "Passed", "ok": "Passed", "✓": "Passed", "đạt": "Passed",
    "fail": "Failed", "failed": "Failed", "failed!": "Failed", "lỗi": "Failed", "không đạt": "Failed",
    "block": "Blocked", "blocked": "Blocked", "bị block": "Blocked",
    "retest": "Retest", "re-test": "Retest", "test lại": "Retest",
    "skip": "Skipped", "skipped": "Skipped", "bỏ qua": "Skipped",
    "n/a": "Skipped",
    "not run": "Not Run", "notrun": "Not Run", "chưa test": "Not Run",
    "": "Not Run",
}

PRIORITY_MAP = {
    "p1": "Critical", "critical": "Critical", "blocker": "Critical",
    "p2": "High", "high": "High", "cao": "High",
    "p3": "Medium", "medium": "Medium", "med": "Medium", "trung bình": "Medium",
    "p4": "Low", "low": "Low", "thấp": "Low",
}

SEVERITY_MAP = {
    "critical": "Critical", "blocker": "Critical",
    "major": "Major", "high": "Major",
    "minor": "Minor", "medium": "Minor",
    "trivial": "Trivial", "low": "Trivial",
}


def _normalize_column(name: str) -> Optional[str]:
    """Map a raw column name to canonical field name."""
    key = str(name).strip().lower()
    return COLUMN_MAP.get(key)


def _normalize_status(val: str) -> str:
    return STATUS_MAP.get(str(val).strip().lower(), "Not Run")


def _normalize_priority(val: str) -> str:
    return PRIORITY_MAP.get(str(val).strip().lower(), "Medium")


def _normalize_severity(val: str) -> str:
    return SEVERITY_MAP.get(str(val).strip().lower(), "Minor")


def _normalize_bool(val) -> bool:
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() in ("yes", "true", "1", "x", "✓", "có", "đã fix", "fixed")


def _find_header_row_openpyxl(file_bytes: bytes) -> tuple[int, list[str]]:
    """Use openpyxl to find the header row — handles merged cells correctly."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    best_row = 0
    best_cols: list[str] = []
    best_score = -1

    for row_idx in range(min(10, ws.max_row)):
        row = ws[row_idx + 1]
        cols = [str(cell.value or "").strip() for cell in row]
        score = sum(1 for c in cols if _normalize_column(c) is not None)
        non_empty = sum(1 for c in cols if c)
        if score > best_score and non_empty >= 2:
            best_score = score
            best_row = row_idx
            best_cols = cols

    wb.close()
    return best_row, best_cols


def parse_file(
    file_bytes: bytes,
    filename: str,
    suite_name: str,
) -> dict:
    """
    Parse Excel or CSV file into structured test cases.
    Returns: { cases: [...], warnings: [...], column_map: {...}, errors: [...] }
    """
    ext = filename.rsplit(".", 1)[-1].lower()
    warnings = []
    errors = []

    try:
        if ext in ("xlsx", "xls"):
            # Use openpyxl to find correct header row (handles merged cells)
            header_row_idx, detected_cols = _find_header_row_openpyxl(file_bytes)
            if header_row_idx > 0:
                warnings.append(f"Header detected at row {header_row_idx + 1} (skipped {header_row_idx} title row(s) above).")
            df = pd.read_excel(io.BytesIO(file_bytes), dtype=str, header=header_row_idx)
        elif ext == "csv":
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
        else:
            return {"cases": [], "warnings": [], "errors": [f"Unsupported file type: .{ext}. Only .xlsx and .csv are supported."], "column_map": {}}
    except Exception as e:
        return {"cases": [], "warnings": [], "errors": [f"Failed to read file: {e}"], "column_map": {}}

    # Fill NaN with empty string
    df = df.fillna("")
    raw_cols = list(df.columns)

    # Map columns
    col_mapping = {}  # raw_col -> canonical
    for col in raw_cols:
        canonical = _normalize_column(col)
        if canonical:
            col_mapping[col] = canonical
        else:
            warnings.append(f"Column '{col}' was not recognized and will be ignored.")

    if "description" not in col_mapping.values():
        errors.append("Required column 'Test Case Description' (or similar) not found. Cannot import.")
        return {"cases": [], "warnings": warnings, "errors": errors, "column_map": col_mapping}

    cases = []
    tc_ids_seen = set()

    for row_idx, row in df.iterrows():
        case: dict = {
            "id": str(uuid.uuid4()),
            "tc_id": None,
            "module": None,
            "feature": None,
            "requirement_id": None,
            "test_type": "Functional",
            "priority": "Medium",
            "severity": "Minor",
            "automation_status": "Manual",
            "precondition": None,
            "description": "",
            "steps": None,
            "test_data": None,
            "expected_result": None,
            "actual_result": None,
            "status": "Not Run",
            "bug_id": None,
            "environment": None,
            "browser": None,
            "platform": None,
            "executed_by": None,
            "execution_date": None,
            "dev_owner": None,
            "dev_fixed": False,
            "notes": None,
            "tags": [],
            "row_order": int(row_idx),
        }

        for raw_col, canonical in col_mapping.items():
            val = str(row.get(raw_col, "")).strip()
            if not val:
                continue

            if canonical == "status":
                case["status"] = _normalize_status(val)
            elif canonical == "priority":
                case["priority"] = _normalize_priority(val)
            elif canonical == "severity":
                case["severity"] = _normalize_severity(val)
            elif canonical == "dev_fixed":
                case["dev_fixed"] = _normalize_bool(val)
            elif canonical == "tags":
                case["tags"] = [t.strip() for t in val.split(",") if t.strip()]
            else:
                case[canonical] = val

        # Skip completely empty rows
        if not case["description"] and not case["tc_id"]:
            continue

        # Skip group/section header rows (e.g. "▼ A. Logic cắt chuỗi cơ bản")
        desc = case["description"] or ""
        is_group_row = (
            desc.startswith("▼") or desc.startswith("▶") or
            desc.startswith("■") or desc.startswith("→") or
            desc.startswith("#") or desc.startswith("===") or
            desc.upper() == desc and len(desc) > 5 and not case["tc_id"]  # ALL CAPS section header
        )
        if is_group_row:
            continue

        # Warn on duplicate TC_ID
        if case["tc_id"]:
            if case["tc_id"] in tc_ids_seen:
                warnings.append(f"Row {row_idx+2}: Duplicate TC_ID '{case['tc_id']}' — kept but flagged.")
            tc_ids_seen.add(case["tc_id"])
        else:
            case["tc_id"] = f"TC-{row_idx+1:03d}"

        # Warn on missing expected result (only for real test cases, not group headers)
        if not case["expected_result"] and not desc.startswith("▼") and not desc.startswith("▶"):
            warnings.append(f"Row {row_idx+2} ({case['tc_id']}): Missing 'Expected Result' — consider adding it for better test quality.")

        cases.append(case)

    logger.info("excel_parsed", filename=filename, total=len(cases), warnings=len(warnings))
    return {
        "cases": cases,
        "warnings": warnings,
        "errors": errors,
        "column_map": {k: v for k, v in col_mapping.items()},
        "total": len(cases),
    }


def export_to_excel(cases: list[dict], suite_name: str) -> bytes:
    """Export test cases back to Excel."""
    export_cols = [
        "tc_id", "module", "feature", "test_type", "priority", "severity",
        "precondition", "description", "steps", "test_data", "expected_result",
        "actual_result", "status", "bug_id", "environment", "browser",
        "executed_by", "execution_date", "dev_fixed", "notes", "tags",
    ]
    display_names = {
        "tc_id": "TC_ID", "module": "Module", "feature": "Feature",
        "test_type": "Test Type", "priority": "Priority", "severity": "Severity",
        "precondition": "Precondition", "description": "Test Case Description",
        "steps": "Steps / Actions", "test_data": "Test Data",
        "expected_result": "Expected Result", "actual_result": "Actual Result",
        "status": "Status", "bug_id": "Bug ID", "environment": "Environment",
        "browser": "Browser", "executed_by": "Executed By",
        "execution_date": "Execution Date", "dev_fixed": "Dev Fixed",
        "notes": "Notes", "tags": "Tags",
    }

    rows = []
    for c in cases:
        row = {}
        for col in export_cols:
            val = c.get(col, "")
            if col == "tags" and isinstance(val, list):
                val = ", ".join(val)
            elif col == "dev_fixed":
                val = "Yes" if val else "No"
            row[display_names.get(col, col)] = val or ""
        rows.append(row)

    df = pd.DataFrame(rows)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name=suite_name[:31], index=False)
        wb = writer.book
        ws = writer.sheets[suite_name[:31]]

        # Header format
        hdr_fmt = wb.add_format({"bold": True, "bg_color": "#6d28d9", "font_color": "white", "border": 1})
        for col_num, col_name in enumerate(df.columns):
            ws.write(0, col_num, col_name, hdr_fmt)
            ws.set_column(col_num, col_num, max(15, len(col_name) + 2))

    return buf.getvalue()
